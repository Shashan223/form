from django.shortcuts import render
from django.conf import settings
import os, uuid
from django.contrib import messages


from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import io
import zipfile
from django.conf import settings
import os

# views.py
from django.shortcuts import render, redirect
from customapp.models import CustomerRegistration
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
import os
from django.contrib.auth import authenticate, login
from django.shortcuts import render, get_object_or_404


def dashboard(request):
    customers = CustomerRegistration.objects.all()
    return render(request, 'customTemp/dashboard.html', {'customers': customers})

def save_file(file_field_name, upload_subdirectory, request):
    file = request.FILES.get(file_field_name)
    if file:
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, upload_subdirectory))
        filename = fs.save(file.name, file)
        return os.path.join(upload_subdirectory, filename)
    return None

def clean_field(value, field_type=int):
    if value in (None, ''):
        return None
    try:
        return field_type(value)
    except (ValueError, TypeError):
        return None


def register_customer(request):
    if request.method == 'POST':
        
        
        
        fs_customer_registration = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'customer_registration'))
        fs_gst_certificate = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'gst_certificates'))
        fs_pan_card = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'pan_cards'))
        fs_balance_sheet = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'balance_sheets'))
        fs_company_certificate = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'company_certificates'))
        fs_bank_statement = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'bank_statements'))
        fs_audited_financials = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'audited_financials'))
        fs_kyc_document = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'kyc_documents'))
        fs_security_cheques = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'security_cheques'))

    # Collect form data 
        # Collecting form data
        data = {
            
            'company_name': request.POST.get('company_name'),
            'registered_address': request.POST.get('registered_address'),
            'pin_code': request.POST.get('pin_code'),
            'landline_no': request.POST.get('landline_no'),
            'telefax_no': request.POST.get('telefax_no'),
            'finance_head_name': request.POST.get('finance_head_name'),
            'finance_head_email': request.POST.get('finance_head_email'),
            'finance_head_mobile': request.POST.get('finance_head_mobile'),
            'billing_address1': request.POST.get('billing_address1'),
            'shipping_address1': request.POST.get('shipping_address1'),
            'billing_contact_name': request.POST.get('billing_contact_name'),
            'shipping_contact_name': request.POST.get('shipping_contact_name'),
            'billing_contact_mobile': request.POST.get('billing_contact_mobile'),
            'shipping_contact_mobile': request.POST.get('shipping_contact_mobile'),
            'billing_contact_email': request.POST.get('billing_contact_email'),
            'shipping_contact_email': request.POST.get('shipping_contact_email'),
            'pan_card_no': request.POST.get('pan_card_no'),
            'gst_no': request.POST.get('gst_no'),
            'msme_no': request.POST.get('msme_no'),
            'cin_no': request.POST.get('cin_no'),
            'company_type': request.POST.getlist('company_type'),
            'business_nature': request.POST.getlist('business_nature'),
            'office_space': request.POST.get('office_space'),
            'office_space_area': request.POST.get('office_space_area'),
            'business_start_year': clean_field(request.POST.get('business_start_year'), int),
            'net_worth': request.POST.get('net_worth'),
            'equity_capital': request.POST.get('equity_capital'),
            'turnover_2023': request.POST.get('turnover_2023'),
            'turnover_2022': request.POST.get('turnover_2022'),
            'turnover_2021': request.POST.get('turnover_2021'),
            'no_of_employees': clean_field(request.POST.get('no_of_employees'), int),

            'no_of_branches': clean_field(request.POST.get('no_of_branches'), int),
            'director_name1': request.POST.get('director_name1'),
            'director_name2': request.POST.get('director_name2'),
            'director_address1': request.POST.get('director_address1'),
            'director_email1': request.POST.get('director_email1'),
            'director_mobile1': request.POST.get('director_mobile1'),
            'bank_name': request.POST.get('bank_name'),
            'bank_address1': request.POST.get('bank_address1'),
            'bank_branch_name': request.POST.get('bank_branch_name'),
            'bank_signatories': clean_field(request.POST.get('bank_signatories'), int),
            'bank_account_no': request.POST.get('bank_account_no'),
            'bank_ifsc_code': request.POST.get('bank_ifsc_code'),
            'account_type': request.POST.getlist('account_type'),
            'bank_limit': request.POST.get('bank_limit'),
            'cash_credit_limit': request.POST.get('cash_credit_limit'),
            'overdraft_limit': request.POST.get('overdraft_limit'),
            'other_limit': request.POST.get('other_limit'),
            'ref_name1': request.POST.get('ref_name1'),
            'ref_address1_1': request.POST.get('ref_address1_1'),
            'ref_telephone1': request.POST.get('ref_telephone1'),
            'ref_contact_person1': request.POST.get('ref_contact_person1'),
            'ref_mobile1': request.POST.get('ref_mobile1'),
            'ref_email1': request.POST.get('ref_email1'),
            'ref_credit_period1': request.POST.get('ref_credit_period1'),
            'ref_credit_limit1': request.POST.get('ref_credit_limit1'),
            'signed_by': request.POST.get('signed_by'),
            'designation': request.POST.get('designation'),
            'office_ref_person': request.POST.get('office_ref_person'),
            'office_branch': request.POST.get('office_branch'),
            'office_credit_amount': request.POST.get('office_credit_amount'),
            'office_credit_days': request.POST.get('office_credit_days'),
            'office_product': request.POST.get('office_product'),
            'office_date': request.POST.get('office_date'),
            'office_authorise_by': request.POST.get('office_authorise_by'),
            'office_approved_amount': request.POST.get('office_approved_amount'),
            'remark': request.POST.get('office_remark'),
  
           
        }
        try:
            customer_registration_path = save_file('doc_customer_registration', 'customer_registration', request)
            gst_certificate_path = save_file('doc_gst_certificate', 'gst_certificates', request)
            pan_card_path = save_file('doc_pan_card', 'pan_cards', request)
            balance_sheet_partnership_path = save_file('doc_balance_sheet_partnership', 'balance_sheets', request)
            company_certificate_path = save_file('doc_company_certificate', 'company_certificates', request)
            bank_statement_path = save_file('doc_bank_statement', 'bank_statements', request)
            audited_financials_path = save_file('doc_audited_financials', 'audited_financials', request)
            kyc_document_path = save_file('doc_kyc_document', 'kyc_documents', request)
            security_cheques_path = save_file('doc_security_cheques', 'security_cheques', request)

            customer = CustomerRegistration(**data)
            customer.customer_registration = customer_registration_path
            customer.gst_certificate = gst_certificate_path
            customer.pan_card = pan_card_path
            customer.balance_sheet = balance_sheet_partnership_path
            customer.company_certificate = company_certificate_path
            customer.bank_statement = bank_statement_path
            customer.audited_financials = audited_financials_path
            customer.kyc_document = kyc_document_path
            customer.security_cheques = security_cheques_path
            customer.save()

            saved_files = {
                'Customer Registration': customer_registration_path,
                'Gst Certificate': gst_certificate_path,
                'Pan Card': pan_card_path,
                'Balance Sheet': balance_sheet_partnership_path,
                'Company Certificate': company_certificate_path,
                'Bank Statement': bank_statement_path,
                'Audited Financials': audited_financials_path,
                'Kyc Document': kyc_document_path,
                'Security Cheques': security_cheques_path,
            }

            return render(request, 'customTemp/success.html', {'files': saved_files, 'MEDIA_URL': settings.MEDIA_URL})

        except Exception as e:
            print(f"An error occurred: {e}")
            return render(request, 'customTemp/form.html', {'error': str(e)})

    return render(request, 'customTemp/form.html')
import os
import io
import zipfile
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from .models import CustomerRegistration

def download_customer_data(request, pk):
    customer = get_object_or_404(CustomerRegistration, pk=pk)

    # Create a ZIP file in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Generate PDF of the customer detail page
        html_string = render_to_string('customTemp/customer_detail.html', {'customer': customer})
        html = HTML(string=html_string, base_url=request.build_absolute_uri())

        # Apply A4 layout and custom CSS to remove borders
        css = CSS(string='''
    @page {
        size: A4;
        margin: 1cm;
    }

    html, body {
        margin: 0;
        padding: 0;
        width: 100%;
        font-family: Arial, sans-serif;
        font-size: 12px;
        box-sizing: border-box;
        overflow-wrap: break-word;
    }

    .content-wrapper {
        width: 100%;
        max-width: 170mm;  /* A4 width (210mm) - 2x margin - buffer */
        margin: 0 auto;
        box-sizing: border-box;
    }

    table {
        width: 100%;
        table-layout: fixed;
        border-collapse: collapse;
        word-wrap: break-word;
    }

    td, th {
        word-wrap: break-word;
        overflow-wrap: break-word;
    }

    input, textarea, select {
        border: none !important;
        outline: none !important;
        background: transparent;
        width: 100%;
    }
''')


        pdf_buffer = html.write_pdf(stylesheets=[css])
        zip_file.writestr(f'customer_detail_{customer.company_name or customer.id}.pdf', pdf_buffer)

        # Add attached files to the ZIP file
        file_fields = [
            customer.printer_report,
            customer.gst_certificate,
            customer.pan_card,
            customer.balance_sheet,
            customer.company_certificate,
            customer.bank_statement,
            customer.audited_financials,
            customer.kyc_document,
            customer.security_cheques,
        ]

        for file_field in file_fields:
            if file_field:
                file_path = file_field.path
                file_name = os.path.basename(file_path)
                try:
                    with open(file_path, 'rb') as f:
                        zip_file.writestr(f'attachments/{file_name}', f.read())
                except FileNotFoundError:
                    print(f"Warning: File not found at {file_path}")

    # Prepare the HTTP response
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="customer_data_{customer.company_name or customer.id}.zip"'
    return response
import io
import os
import zipfile
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook



import io
import zipfile
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import get_object_or_404
import os
from django.conf import settings

import io
import os
import zipfile
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook

import io
import zipfile
import os
from openpyxl import Workbook
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
import io
import zipfile
import os
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import CustomerRegistration # Assuming CustomerRegistration is in the same app
def download_customer_dataa(request, pk):
    customer = get_object_or_404(CustomerRegistration, pk=pk)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # --- PDF Generation (Existing Code) ---
        html_string = render_to_string('customTemp/customer_detail.html', {'customer': customer})
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        css = CSS(string='''
    @page {
        size: A4;
        margin: 1cm;
    }

    html, body {
        margin: 0;
        padding: 0;
        width: 100%;
        font-family: Arial, sans-serif;
        font-size: 12px;
        box-sizing: border-box;
        overflow-wrap: break-word;
    }

    .content-wrapper {
        width: 100%;
        max-width: 170mm;
        margin: 0 auto;
        box-sizing: border-box;
    }

    table {
        width: 100%;
        table-layout: fixed;
        border-collapse: collapse;
        word-wrap: break-word;
    }

    td, th {
        word-wrap: break-word;
        overflow-wrap: break-word;
    }

    input, textarea, select {
        border: none !important;
        outline: none !important;
        background: transparent;
        width: 100%;
    }

    /* Hide default checkboxes/radios */
    input[type="checkbox"],
    input[type="radio"] {
        appearance: none;
        -webkit-appearance: none;
        -moz-appearance: none;
        background-color: white;
        border: 1px solid white;
        width: 10px;
        height: 10px;
        display: inline-block;
        position: relative;
        margin: 0 4px 0 0;
        vertical-align: middle;
    }

    /* Checked checkbox */
    input[type="checkbox"]:checked::before {
        content: "✓";
        color: white;
        font-size: 12px;
        position: absolute;
        left: 2px;
        top: -1px;
    }

    /* Radio as circle */
    input[type="radio"] {
        border-radius: 50%;
    }

    /* Checked radio circle inside */
    input[type="radio"]:checked::before {
        content: "";
        width: 8px;
        height: 8px;
        background: white;
        border-radius: 50%;
        position: absolute;
        top: 3px;
        left: 3px;
    }

    .checkbox-label, .radio-label {
        display: inline-flex;
        align-items: center;
        gap: 6px;
    }
''')
        pdf_buffer = html.write_pdf(stylesheets=[css]) 
        zip_file.writestr(f'customer_detail_{customer.company_name or customer.id}.pdf', pdf_buffer)

        # --- Excel Generation (New/Modified Code) ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Registration"

        # Set default font for the worksheet
        ws.sheet_properties.tabColor = "1072BA"
        
        # Set column widths to match the template
        column_widths = {
            'A': 5, 'B': 5, 'C': 40, 'D': 5, 'E': 15, 'F': 5, 'G': 15, 
            'H': 5, 'I': 15, 'J': 5, 'K': 15, 'L': 5, 'M': 15, 'N': 5, 'O': 15
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Header Section
        ws.merge_cells('C1:O1')
        header_cell = ws['C1']
        header_cell.value = "RITCOM SYSTEMS AND SERVICES PRIVATE LIMITED"
        header_cell.font = Font(name='Arial', size=12, bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('C2:O2')
        address_cell = ws['C2']
        address_cell.value = "Unit No 2 , Aristocrate, Lajya Compound, Mogra Road, Andheri(East) Mumbai 400069."
        address_cell.font = Font(name='Arial', size=10)
        address_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Customer Registration Form Title
        ws.merge_cells('A4:O4')
        title_cell = ws['A4']
        title_cell.value = "Customer Registration Form"
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Company Information Section
        ws['A6'] = "Company Name:-"
        ws['A6'].font = Font(name='Arial', size=10, bold=True)
        ws.merge_cells('C6:O6')
        ws['C6'] = customer.company_name or ""
        ws['C6'].font = Font(name='Arial', size=10)
        ws['C6'].alignment = Alignment(wrap_text=True)

        ws['A7'] = "Registered Address:-"
        ws['A7'].font = Font(name='Arial', size=10, bold=True)
        ws.merge_cells('C7:O7')
        ws['C7'] = customer.registered_address or ""
        ws['C7'].font = Font(name='Arial', size=10)
        ws['C7'].alignment = Alignment(wrap_text=True)

        ws['N8'] = "Pin code:-"
        ws['N8'].font = Font(name='Arial', size=10, bold=True)
        ws['O8'] = customer.pin_code or ""
        ws['O8'].font = Font(name='Arial', size=10)

        # Contact Information
        ws['A9'] = "Land Line No:-"
        ws['A9'].font = Font(name='Arial', size=10, bold=True)
        ws['C9'] = customer.landline_no or ""
        ws['C9'].font = Font(name='Arial', size=10)

        ws['M9'] = "Tele Fax No:-"
        ws['M9'].font = Font(name='Arial', size=10, bold=True)
        ws['O9'] = customer.telefax_no or ""
        ws['O9'].font = Font(name='Arial', size=10)

        ws['A10'] = "Name of Finance Head:-"
        ws['A10'].font = Font(name='Arial', size=10, bold=True)
        ws['C10'] = customer.finance_head_name or ""
        ws['C10'].font = Font(name='Arial', size=10)

        ws['A11'] = "Email ID of Finance Head:-"
        ws['A11'].font = Font(name='Arial', size=10, bold=True)
        ws['C11'] = customer.finance_head_email or ""
        ws['C11'].font = Font(name='Arial', size=10)

        ws['M11'] = "Mobile No:-"
        ws['M11'].font = Font(name='Arial', size=10, bold=True)
        ws['O11'] = customer.finance_head_mobile or ""
        ws['O11'].font = Font(name='Arial', size=10)

        # Billing & Shipping Address
        ws.merge_cells('A12:O12')
        ws['A12'] = "Billing & Shipping Address(Multiple Godown Address)"
        ws['A12'].font = Font(name='Arial', size=10, bold=True)

        ws['E13'] = "Billing Address / Office Address"
        ws['E13'].font = Font(name='Arial', size=10, bold=True)
        ws['K13'] = "Shipping Address / Godown Address"
        ws['K13'].font = Font(name='Arial', size=10, bold=True)

        ws.merge_cells('E14:I14')
        ws['E14'] = customer.billing_address1 or ""
        ws['E14'].font = Font(name='Arial', size=10)
        ws['E14'].alignment = Alignment(wrap_text=True)

        ws.merge_cells('K14:O14')
        ws['K14'] = customer.shipping_address1 or ""
        ws['K14'].font = Font(name='Arial', size=10)
        ws['K14'].alignment = Alignment(wrap_text=True)

        ws['A15'] = "Contact Person Name:-"
        ws['A15'].font = Font(name='Arial', size=10, bold=True)
        ws['E15'] = customer.billing_contact_name or ""
        ws['E15'].font = Font(name='Arial', size=10)
        ws['K15'] = customer.shipping_contact_name or ""
        ws['K15'].font = Font(name='Arial', size=10)

        ws['A16'] = "Contact Person Mobile No:-"
        ws['A16'].font = Font(name='Arial', size=10, bold=True)
        ws['E16'] = customer.billing_contact_mobile or ""
        ws['E16'].font = Font(name='Arial', size=10)
        ws['K16'] = customer.shipping_contact_mobile or ""
        ws['K16'].font = Font(name='Arial', size=10)

        ws['A17'] = "Contact Person E-Mail ID:-"
        ws['A17'].font = Font(name='Arial', size=10, bold=True)
        ws['E17'] = customer.billing_contact_email or ""
        ws['E17'].font = Font(name='Arial', size=10)
        ws['K17'] = customer.shipping_contact_email or ""
        ws['K17'].font = Font(name='Arial', size=10)

        # Tax Registration Details
        ws.merge_cells('A18:O18')
        ws['A18'] = "Tax Registration Details"
        ws['A18'].font = Font(name='Arial', size=10, bold=True)

        ws['A19'] = "PAN CARD No:-"
        ws['A19'].font = Font(name='Arial', size=10, bold=True)
        ws['C19'] = customer.pan_card_no or ""
        ws['C19'].font = Font(name='Arial', size=10)

        ws['I19'] = "GST No:-"
        ws['I19'].font = Font(name='Arial', size=10, bold=True)
        ws['K19'] = customer.gst_no or ""
        ws['K19'].font = Font(name='Arial', size=10)

        ws['A20'] = "MSME No:-"
        ws['A20'].font = Font(name='Arial', size=10, bold=True)
        ws['C20'] = customer.msme_no or ""
        ws['C20'].font = Font(name='Arial', size=10)

        ws['I20'] = "CIN No:-"
        ws['I20'].font = Font(name='Arial', size=10, bold=True)
        ws['K20'] = customer.cin_no or ""
        ws['K20'].font = Font(name='Arial', size=10)

        # Type of Company
        ws['A21'] = "Type Of Company:-"
        ws['A21'].font = Font(name='Arial', size=10, bold=True)
        
        # Set checkboxes based on company_type
        company_type = customer.company_type or ""
        ws['E21'] = "Proprietor:-"
        ws['E21'].font = Font(name='Arial', size=10, bold=True)
        ws['F21'] = "✓" if "Proprietor" in company_type else ""
        ws['F21'].font = Font(name='Arial', size=10)
        
        ws['H21'] = "Partnership:-"
        ws['H21'].font = Font(name='Arial', size=10, bold=True)
        ws['I21'] = "✓" if "Partnership" in company_type else ""
        ws['I21'].font = Font(name='Arial', size=10)
        
        ws['K21'] = "Pvt Ltd:-"
        ws['K21'].font = Font(name='Arial', size=10, bold=True)
        ws['L21'] = "✓" if "Pvt Ltd" in company_type else ""
        ws['L21'].font = Font(name='Arial', size=10)
        
        ws['N21'] = "Ltd:-"
        ws['N21'].font = Font(name='Arial', size=10, bold=True)
        ws['O21'] = "✓" if "Ltd" in company_type else ""
        ws['O21'].font = Font(name='Arial', size=10)

        # Nature of Business
        ws['A22'] = "Nature of Business:-"
        ws['A22'].font = Font(name='Arial', size=10, bold=True)
        
        business_nature = customer.business_nature or ""
        ws['E22'] = "System Integrator:-"
        ws['E22'].font = Font(name='Arial', size=10, bold=True)
        ws['F22'] = "✓" if "System Integrator" in business_nature else ""
        ws['F22'].font = Font(name='Arial', size=10)
        
        ws['H22'] = "Reseller:-"
        ws['H22'].font = Font(name='Arial', size=10, bold=True)
        ws['I22'] = "✓" if "Reseller" in business_nature else ""
        ws['I22'].font = Font(name='Arial', size=10)
        
        ws['K22'] = "Retailer:-"
        ws['K22'].font = Font(name='Arial', size=10, bold=True)
        ws['L22'] = "✓" if "Retailer" in business_nature else ""
        ws['L22'].font = Font(name='Arial', size=10)
        
        ws['N22'] = "E-Retailer:-"
        ws['N22'].font = Font(name='Arial', size=10, bold=True)
        ws['O22'] = "✓" if "E-Retailer" in business_nature else ""
        ws['O22'].font = Font(name='Arial', size=10)

        # Office Space
        ws['A23'] = "Office Space:-"
        ws['A23'].font = Font(name='Arial', size=10, bold=True)
        
        office_space = customer.office_space or ""
        ws['E23'] = "Rented:-"
        ws['E23'].font = Font(name='Arial', size=10, bold=True)
        ws['F23'] = "✓" if "Rented" in office_space else ""
        ws['F23'].font = Font(name='Arial', size=10)
        
        ws['H23'] = "Ownership:-"
        ws['H23'].font = Font(name='Arial', size=10, bold=True)
        ws['I23'] = "✓" if "Ownership" in office_space else ""
        ws['I23'].font = Font(name='Arial', size=10)
        
        ws['K23'] = "Office Space:-"
        ws['K23'].font = Font(name='Arial', size=10, bold=True)
        ws['L23'] = customer.office_space_area or ""
        ws['L23'].font = Font(name='Arial', size=10)

        # Company Turnover
        ws.merge_cells('A24:O24')
        ws['A24'] = "COMPANY TURNOVER (FOR LAST THREE YEARS)"
        ws['A24'].font = Font(name='Arial', size=10, bold=True)

        ws['A25'] = "Business Start Year:-"
        ws['A25'].font = Font(name='Arial', size=10, bold=True)
        ws['C25'] = str(customer.business_start_year) if customer.business_start_year else ""
        ws['C25'].font = Font(name='Arial', size=10)

        ws['G25'] = "Net Worth:-"
        ws['G25'].font = Font(name='Arial', size=10, bold=True)
        ws['I25'] = customer.net_worth or ""
        ws['I25'].font = Font(name='Arial', size=10)

        ws['M25'] = "Equity/Capital :-"
        ws['M25'].font = Font(name='Arial', size=10, bold=True)
        ws['O25'] = customer.equity_capital or ""
        ws['O25'].font = Font(name='Arial', size=10)

        ws['C26'] = "Year 2022 - 2023"
        ws['C26'].font = Font(name='Arial', size=10, bold=True)
        ws['G26'] = "Year 2021 - 2022"
        ws['G26'].font = Font(name='Arial', size=10, bold=True)
        ws['K26'] = "Year 2020 - 2021"
        ws['K26'].font = Font(name='Arial', size=10, bold=True)

        ws['C27'] = customer.turnover_2023 or ""
        ws['C27'].font = Font(name='Arial', size=10)
        ws['G27'] = customer.turnover_2022 or ""
        ws['G27'].font = Font(name='Arial', size=10)
        ws['K27'] = customer.turnover_2021 or ""
        ws['K27'].font = Font(name='Arial', size=10)

        ws['M26'] = "No Of Employees:-"
        ws['M26'].font = Font(name='Arial', size=10, bold=True)
        ws['O26'] = str(customer.no_of_employees) if customer.no_of_employees else ""
        ws['O26'].font = Font(name='Arial', size=10)

        ws['M27'] = "No of Branches:-"
        ws['M27'].font = Font(name='Arial', size=10, bold=True)
        ws['O27'] = str(customer.no_of_branches) if customer.no_of_branches else ""
        ws['O27'].font = Font(name='Arial', size=10)

        # Proprietor/Partners/Directors details
        ws.merge_cells('A28:O28')
        ws['A28'] = "Proprietor/ Partners/ Directors details:"
        ws['A28'].font = Font(name='Arial', size=10, bold=True)

        ws['A29'] = "Name:-"
        ws['A29'].font = Font(name='Arial', size=10, bold=True)
        ws['C29'] = customer.director_name1 or ""
        ws['C29'].font = Font(name='Arial', size=10)

        ws['A30'] = "Residence Address:-"
        ws['A30'].font = Font(name='Arial', size=10, bold=True)
        ws.merge_cells('C30:O30')
        ws['C30'] = customer.director_address1 or ""
        ws['C30'].font = Font(name='Arial', size=10)
        ws['C30'].alignment = Alignment(wrap_text=True)

        ws['A32'] = "E-Mail ID"
        ws['A32'].font = Font(name='Arial', size=10, bold=True)
        ws['C32'] = customer.director_email1 or ""
        ws['C32'].font = Font(name='Arial', size=10)

        ws['A33'] = "Mobile No:-"
        ws['A33'].font = Font(name='Arial', size=10, bold=True)
        ws['C33'] = customer.director_mobile1 or ""
        ws['C33'].font = Font(name='Arial', size=10)

        # Banker Information
        ws.merge_cells('A34:O34')
        ws['A34'] = "BANKER INFORMATION"
        ws['A34'].font = Font(name='Arial', size=10, bold=True)

        ws['A35'] = "Bank Name:-"
        ws['A35'].font = Font(name='Arial', size=10, bold=True)
        ws['C35'] = customer.bank_name or ""
        ws['C35'].font = Font(name='Arial', size=10)

        ws['A36'] = "Bank Address:-"
        ws['A36'].font = Font(name='Arial', size=10, bold=True)
        ws.merge_cells('C36:O36')
        ws['C36'] = customer.bank_address1 or ""
        ws['C36'].font = Font(name='Arial', size=10)
        ws['C36'].alignment = Alignment(wrap_text=True)

        ws['A38'] = "Branch Name:-"
        ws['A38'].font = Font(name='Arial', size=10, bold=True)
        ws['C38'] = customer.bank_branch_name or ""
        ws['C38'].font = Font(name='Arial', size=10)

        ws['M38'] = "No. of Signatories :-"
        ws['M38'].font = Font(name='Arial', size=10, bold=True)
        ws['O38'] = str(customer.bank_signatories) if customer.bank_signatories else ""
        ws['O38'].font = Font(name='Arial', size=10)

        ws['A39'] = "Account No:-"
        ws['A39'].font = Font(name='Arial', size=10, bold=True)
        ws['C39'] = customer.bank_account_no or ""
        ws['C39'].font = Font(name='Arial', size=10)

        ws['K39'] = "IFSC / NEFT Code:-"
        ws['K39'].font = Font(name='Arial', size=10, bold=True)
        ws['M39'] = customer.bank_ifsc_code or ""
        ws['M39'].font = Font(name='Arial', size=10)

        # Type of Account
        ws['A40'] = "Type of Account:-"
        ws['A40'].font = Font(name='Arial', size=10, bold=True)
        
        account_type = customer.account_type or ""
        ws['E40'] = "Current:-"
        ws['E40'].font = Font(name='Arial', size=10, bold=True)
        ws['F40'] = "✓" if "Current" in account_type else ""
        ws['F40'].font = Font(name='Arial', size=10)
        
        ws['H40'] = "Cash Credit:-"
        ws['H40'].font = Font(name='Arial', size=10, bold=True)
        ws['I40'] = "✓" if "Cash Credit" in account_type else ""
        ws['I40'].font = Font(name='Arial', size=10)
        
        ws['K40'] = "Overdraft:-"
        ws['K40'].font = Font(name='Arial', size=10, bold=True)
        ws['L40'] = "✓" if "Overdraft" in account_type else ""
        ws['L40'].font = Font(name='Arial', size=10)
        
        ws['N40'] = "Others:-"
        ws['N40'].font = Font(name='Arial', size=10, bold=True)
        ws['O40'] = "✓" if account_type and "Current" not in account_type and "Cash Credit" not in account_type and "Overdraft" not in account_type else ""
        ws['O40'].font = Font(name='Arial', size=10)

        # Bank Limits
        ws['A41'] = "Bank Limit:-"
        ws['A41'].font = Font(name='Arial', size=10, bold=True)
        ws['E41'] = customer.bank_limit or ""
        ws['E41'].font = Font(name='Arial', size=10)
        
        ws['H41'] = "Cash Credit:-"
        ws['H41'].font = Font(name='Arial', size=10, bold=True)
        ws['I41'] = customer.cash_credit_limit or ""
        ws['I41'].font = Font(name='Arial', size=10)
        
        ws['K41'] = "Overdraft:-"
        ws['K41'].font = Font(name='Arial', size=10, bold=True)
        ws['L41'] = customer.overdraft_limit or ""
        ws['L41'].font = Font(name='Arial', size=10)
        
        ws['N41'] = "Others:-"
        ws['N41'].font = Font(name='Arial', size=10, bold=True)
        ws['O41'] = customer.other_limit or ""
        ws['O41'].font = Font(name='Arial', size=10)

        # Reference of Distributor/Parties
        ws.merge_cells('A42:O42')
        ws['A42'] = "REFERENCE OF DISTRIBUTOR / PARTIES"
        ws['A42'].font = Font(name='Arial', size=10, bold=True)

        ws['A43'] = "Name:-"
        ws['A43'].font = Font(name='Arial', size=10, bold=True)
        ws['C43'] = customer.ref_name1 or ""
        ws['C43'].font = Font(name='Arial', size=10)

        ws['A44'] = "Address:-"
        ws['A44'].font = Font(name='Arial', size=10, bold=True)
        ws.merge_cells('C44:O44')
        ws['C44'] = customer.ref_address1_1 or ""
        ws['C44'].font = Font(name='Arial', size=10)
        ws['C44'].alignment = Alignment(wrap_text=True)

        ws['A47'] = "Telephone No:-"
        ws['A47'].font = Font(name='Arial', size=10, bold=True)
        ws['C47'] = customer.ref_telephone1 or ""
        ws['C47'].font = Font(name='Arial', size=10)

        ws['E47'] = "Contact Person:-"
        ws['E47'].font = Font(name='Arial', size=10, bold=True)
        ws['G47'] = customer.ref_contact_person1 or ""
        ws['G47'].font = Font(name='Arial', size=10)

        ws['I47'] = "Mobile No:-"
        ws['I47'].font = Font(name='Arial', size=10, bold=True)
        ws['K47'] = customer.ref_mobile1 or ""
        ws['K47'].font = Font(name='Arial', size=10)

        ws['M47'] = "E-Mail ID:-"
        ws['M47'].font = Font(name='Arial', size=10, bold=True)
        ws['O47'] = customer.ref_email1 or ""
        ws['O47'].font = Font(name='Arial', size=10)

        ws['A48'] = "Credit Period:-"
        ws['A48'].font = Font(name='Arial', size=10, bold=True)
        ws['C48'] = customer.ref_credit_period1 or ""
        ws['C48'].font = Font(name='Arial', size=10)

        ws['E48'] = "Credit Limit:-"
        ws['E48'].font = Font(name='Arial', size=10, bold=True)
        ws['G48'] = customer.ref_credit_limit1 or ""
        ws['G48'].font = Font(name='Arial', size=10)

        # Terms of Trade
        ws.merge_cells('A49:O49')
        ws['A49'] = "List of Terms Of Trade Acceptance to be signed for SMOOTH BUSINESS"
        ws['A49'].font = Font(name='Arial', size=10, bold=True)

        terms = [
            "1. For All the transactions P.O. is mandatory to clearly mentioning the SKU,QUANTITY,Applicable GST and Payment Terms.",
            "2. All P.Os /Acknowledged  clearly mentioning billing instruction by you will be binding upon you and under no circumstances any SALES RETURN done.",
            "3. Payment terms will be from date of billing. Any delay in Payment , interest will be applicable",
            "4.Finanace cost debit note payment has to be cleared within 15 days from the date of Debit note",
            "5. Any SUPPORT from VENDOR to be disbursed through DCICL will not be deducted from regular payment till CN is approved by VENDOR/DISTRIBUTOR",
            "6.Billing will be on HOLD unless and until the overdue is cleared along with Delay Interest",
            "7. Any unfulfilled /Pending VENDOR Commitment verbal or written, DCICL will have no role in the same and not responsible for any commitment by VENDOR/DISTRIBUTOR",
            "8. CDC Discount applicable will be product specific and shared at the time of placing the order.",
            "9. CDC payments to be made against delivery OR along with P.O."
        ]

        for i, term in enumerate(terms, start=50):
            ws.merge_cells(f'A{i}:O{i}')
            ws[f'A{i}'] = term
            ws[f'A{i}'].font = Font(name='Arial', size=10)
            ws[f'A{i}'].alignment = Alignment(wrap_text=True)

        # Signature Section
        signature_row = 50 + len(terms)
        ws.merge_cells(f'A{signature_row}:G{signature_row}')
        ws[f'A{signature_row}'] = "Signature of Owner / Partner / Director."
        ws[f'A{signature_row}'].font = Font(name='Arial', size=10, bold=True)

        ws.merge_cells(f'N{signature_row}:O{signature_row}')
        ws[f'N{signature_row}'] = "Company Seal"
        ws[f'N{signature_row}'].font = Font(name='Arial', size=10, bold=True)

        ws[f'A{signature_row + 1}'] = "Signed By:-"
        ws[f'A{signature_row + 1}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'C{signature_row + 1}'] = customer.signed_by or ""
        ws[f'C{signature_row + 1}'].font = Font(name='Arial', size=10)

        ws[f'A{signature_row + 3}'] = "Designation:-"
        ws[f'A{signature_row + 3}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'C{signature_row + 3}'] = customer.designation or ""
        ws[f'C{signature_row + 3}'].font = Font(name='Arial', size=10)

        # Documents to Enclose
        docs_row = signature_row + 5
        ws.merge_cells(f'A{docs_row}:O{docs_row}')
        ws[f'A{docs_row}'] = "Kindly Encl the Below Document."
        ws[f'A{docs_row}'].font = Font(name='Arial', size=10, bold=True)

        documents = [
            "1. Customer Registration Form duly filled & signed",
            "2. GST Certificate (Form 6)",
            "3. Pan Card of the Company",
            "4. For Proprietary / Partnership concerns - Audited Balance Sheet & Partnership Deed",
            "5. For Companies. - Copy of Certificate of Incorporation, Memorandum of Association, Articles of Association",
            "6. Bank Statement for last 3 Months",
            "7. Latest year Audited Balance Sheet / Financial",
            "8. KYC document of Promoter / Director/ Partner / Proprietor (Driving License / Passport / Aadhar Card)",
            "9. 2 Security Cheques."
        ]

        for i, doc in enumerate(documents, start=docs_row + 1):
            ws[f'A{i}'] = doc
            ws[f'A{i}'].font = Font(name='Arial', size=10)

        # Office Use Only
        office_row = docs_row + len(documents) + 1
        ws.merge_cells(f'A{office_row}:O{office_row}')
        ws[f'A{office_row}'] = "For DC Infotech and Communication Limited office use only."
        ws[f'A{office_row}'].font = Font(name='Arial', size=10, bold=True)

        ws[f'A{office_row + 1}'] = "Reference Person:-"
        ws[f'A{office_row + 1}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'C{office_row + 1}'] = customer.office_ref_person or ""
        ws[f'C{office_row + 1}'].font = Font(name='Arial', size=10)

        ws[f'I{office_row + 1}'] = "Branch:-"
        ws[f'I{office_row + 1}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'K{office_row + 1}'] = customer.office_branch or ""
        ws[f'K{office_row + 1}'].font = Font(name='Arial', size=10)

        ws[f'A{office_row + 2}'] = "Credit Amount:-"
        ws[f'A{office_row + 2}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'C{office_row + 2}'] = customer.office_credit_amount or ""
        ws[f'C{office_row + 2}'].font = Font(name='Arial', size=10)

        ws[f'I{office_row + 2}'] = "Credit Days:-"
        ws[f'I{office_row + 2}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'K{office_row + 2}'] = customer.office_credit_days or ""
        ws[f'K{office_row + 2}'].font = Font(name='Arial', size=10)

        ws[f'A{office_row + 3}'] = "Product:-"
        ws[f'A{office_row + 3}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'C{office_row + 3}'] = customer.office_product or ""
        ws[f'C{office_row + 3}'].font = Font(name='Arial', size=10)

        ws[f'I{office_row + 3}'] = "Date:-"
        ws[f'I{office_row + 3}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'K{office_row + 3}'] = customer.office_date or ""
        ws[f'K{office_row + 3}'].font = Font(name='Arial', size=10)

        ws[f'A{office_row + 4}'] = "Authorise By:-"
        ws[f'A{office_row + 4}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'C{office_row + 4}'] = customer.office_authorise_by or ""
        ws[f'C{office_row + 4}'].font = Font(name='Arial', size=10)

        ws[f'I{office_row + 4}'] = "Approved Amount:-"
        ws[f'I{office_row + 4}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'K{office_row + 4}'] = customer.office_approved_amount or ""
        ws[f'K{office_row + 4}'].font = Font(name='Arial', size=10)

        ws[f'A{office_row + 5}'] = "Remark"
        ws[f'A{office_row + 5}'].font = Font(name='Arial', size=10, bold=True)
        ws.merge_cells(f'B{office_row + 5}:O{office_row + 5}')
        ws[f'B{office_row + 5}'] = customer.remark or ""
        ws[f'B{office_row + 5}'].font = Font(name='Arial', size=10)
        ws[f'B{office_row + 5}'].alignment = Alignment(wrap_text=True)

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        zip_file.writestr(f'customer_data_{customer.company_name or customer.id}.xlsx', excel_buffer.getvalue())

        # --- Attach Uploaded Files (Existing Code) ---
        from django.db.models import FileField
        for field in CustomerRegistration._meta.fields:
            if isinstance(field, FileField):
                file_instance = getattr(customer, field.name)
                if file_instance:
                    file_path = file_instance.path
                    file_name = os.path.basename(file_path)
                    try:
                        with open(file_path, 'rb') as f:
                            zip_file.writestr(f'attachments/{file_name}', f.read())
                    except FileNotFoundError:
                        print(f"Warning: File not found at {file_path}")

    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="customer_data_{customer.company_name or customer.id}.zip"'
    return response 

def customer_detail(request, pk):
    customer = get_object_or_404(CustomerRegistration, pk=pk)
    return render(request, 'customTemp/customer_detail.html', {'customer': customer})

def menu(request):
    return render(request,'customTemp/menu.html')
def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if user.is_superuser:
                return redirect('dashboard')  # redirect superusers
            else:
                messages.error(request, 'Invalid username or password.')    # redirect regular users
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'customTemp/login.html')



# Create your views here.
